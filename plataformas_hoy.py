import openpyxl
import pandas as pd
import sqlite3
import re
import os
import streamlit as st
import logging
from datetime import datetime

# Configuración básica de logging
logging.basicConfig(level=logging.INFO, filename='procesamiento.log', filemode='w',
                    format='%(asctime)s - %(levelname)s - %(message)s')

# Ruta predeterminada
default_excel_path = r"C:\Users\capac\OneDrive\Escritorio\Actividades de Sims\bd_sims"

# Mapeos predeterminados basados en el nombre de la pestaña
default_mappings = {
    "WIALON": {
        'Nombre': 'Nombre',
        'Cliente_Cuenta': 'Cuenta',
        'Tipo_de_Dispositivo': 'Tipo de dispositivo',
        'IMEI': 'IMEI',
        'ICCID': 'Iccid',
        'Fecha_de_Activacion': 'Creada',
        'Fecha_de_Desactivacion': 'Desactivación',
        'Hora_de_Ultimo_Mensaje': 'Hora de último mensaje',
        'Ultimo_Reporte': 'Ultimo Reporte',
        'Vehiculo': None,
        'Servicios': None,
        'Grupo': 'Grupos',
        'Telefono': 'Teléfono',
        'Origen': 'WIALON',  # Asignado manualmente
        'Fecha_Archivo': None  # Extraído del nombre del archivo
    },
    "ADAS": {
        'Nombre': 'equipo',
        'Cliente_Cuenta': 'Subordinar',
        'Tipo_de_Dispositivo': 'Modelo',
        'IMEI': 'IMEI',
        'ICCID': 'Iccid',
        'Fecha_de_Activacion': 'Activation Date',
        'Fecha_de_Desactivacion': None,
        'Hora_de_Ultimo_Mensaje': None,
        'Ultimo_Reporte': None,
        'Vehiculo': None,
        'Servicios': None,
        'Grupo': None,
        'Telefono': 'Número de tarjeta SIM',
        'Origen': 'ADAS',  # Asignado manualmente
        'Fecha_Archivo': None  # Extraído del nombre del archivo
    },
    "COMBUSTIBLE": {
        'Nombre': 'Vehículo',
        'Cliente_Cuenta': 'Cuenta',
        'Tipo_de_Dispositivo': 'Tanques',
        'IMEI': None,
        'ICCID': None,
        'Fecha_de_Activacion': None,
        'Fecha_de_Desactivacion': None,
        'Hora_de_Ultimo_Mensaje': None,
        'Ultimo_Reporte': 'Último reporte',
        'Vehiculo': 'Vehículo',
        'Servicios': 'Servicios',
        'Grupo': 'Grupos',
        'Telefono': 'Línea',
        'Origen': 'COMBUSTIBLE',  # Asignado manualmente
        'Fecha_Archivo': None  # Extraído del nombre del archivo
    }
}

def create_database(db_path):
    conn = sqlite3.connect(db_path)
    cursor = conn.cursor()
    cursor.execute(''' 
        CREATE TABLE IF NOT EXISTS datos ( 
            Nombre TEXT,
            Cliente_Cuenta TEXT,
            Tipo_de_Dispositivo TEXT,
            IMEI TEXT,
            ICCID TEXT,
            Fecha_de_Activacion TEXT,
            Fecha_de_Desactivacion TEXT,
            Hora_de_Ultimo_Mensaje TEXT,
            Ultimo_Reporte TEXT,
            Vehiculo TEXT,
            Servicios TEXT,
            Grupo TEXT,
            Telefono TEXT,
            Origen TEXT,
            Fecha_Archivo TEXT,
            UNIQUE(Nombre, Cliente_Cuenta, Telefono)
        ) 
    ''')
    conn.commit()
    conn.close()

# Función para insertar datos en la base de datos con manejo de duplicados
def insert_data(db_path, data):
    conn = sqlite3.connect(db_path)
    cursor = conn.cursor()
    try:
        cursor.executemany(
            '''INSERT OR IGNORE INTO datos (
                Nombre, Cliente_Cuenta, Tipo_de_Dispositivo, IMEI, ICCID,
                Fecha_de_Activacion, Fecha_de_Desactivacion, Hora_de_Ultimo_Mensaje,
                Ultimo_Reporte, Vehiculo, Servicios, Grupo, Telefono, Origen, Fecha_Archivo
            ) VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)''',
            data
        )
        conn.commit()
        logging.info(f"Insertados {cursor.rowcount} registros en la base de datos.")
        inserted = cursor.rowcount
    except sqlite3.IntegrityError as e:
        logging.error(f"Error al insertar datos: {e}")
        inserted = 0
    conn.close()
    return inserted

# Función para limpiar el campo Telefono (sin validación de cantidad de dígitos)
def clean_telefono(telefono):
    if telefono:
        telefono = re.sub(r'\D', '', str(telefono))
        if telefono:  # Solo verifica que no esté vacío después de limpiar
            return telefono
    return None

# Función para extraer la fecha del nombre del archivo
def extract_date_from_filename(filename):
    match = re.search(r'\d{4}-\d{2}-\d{2}', filename)
    if match:
        return match.group(0)
    else:
        return datetime.now().strftime('%Y-%m-%d')

# Función para procesar el archivo Excel con múltiples pestañas
def process_excel_file(excel_file, mappings):
    all_data = []
    invalid_data = []
    total_records = 0
    filename = os.path.basename(excel_file)  # Obtener solo el nombre del archivo
    fecha_archivo = extract_date_from_filename(filename)
    workbook = openpyxl.load_workbook(excel_file, data_only=True)
    for sheet_name in workbook.sheetnames:
        if sheet_name in mappings:
            mapping = mappings[sheet_name]
            sheet = workbook[sheet_name]
            headers = [cell.value for cell in next(sheet.iter_rows(min_row=1, max_row=1))]
            # Crear un diccionario que mapea nombres de columnas a índices
            col_indices = {header: idx for idx, header in enumerate(headers)}
            for row in sheet.iter_rows(min_row=2, values_only=True):
                total_records += 1
                row_dict = {headers[i]: row[i] for i in range(len(headers))}
                record = {}
                is_valid = True
                # Validaciones específicas
                required_field = 'Cliente_Cuenta'  # Solo Cliente_Cuenta es requerido
                column_name = mapping.get(required_field)
                value = row_dict.get(column_name) if column_name else None
                if not value:
                    is_valid = False
                # Si es Cliente_Cuenta, no es necesario verificar otros campos
                if is_valid:
                    # Construcción del registro válido
                    for field in [
                        'Nombre', 'Cliente_Cuenta', 'Tipo_de_Dispositivo', 'IMEI', 'ICCID',
                        'Fecha_de_Activacion', 'Fecha_de_Desactivacion', 'Hora_de_Ultimo_Mensaje',
                        'Ultimo_Reporte', 'Vehiculo', 'Servicios', 'Grupo', 'Telefono', 'Origen', 'Fecha_Archivo'
                    ]:
                        if field == 'Origen':
                            record[field] = mapping['Origen']
                        elif field == 'Fecha_Archivo':
                            record[field] = fecha_archivo
                        else:
                            column_name = mapping.get(field)
                            if column_name:
                                value = row_dict.get(column_name)
                                if field == 'Telefono':
                                    value = clean_telefono(value)
                                record[field] = value
                            else:
                                record[field] = None
                    all_data.append(tuple(record.values()))
                    logging.info(f"Procesado registro válido en pestaña '{sheet_name}': {record}")
                else:
                    invalid_data.append(row_dict)
                    logging.warning(f"Registro inválido en pestaña '{sheet_name}': {row_dict}")
    return all_data, invalid_data, total_records

# Interfaz de usuario con Streamlit
st.title("Carga y Homologación de Datos desde Excel con Múltiples Pestañas")

# Verificar si existe una base de datos previa
today_db_path = os.path.join(default_excel_path, f'{datetime.now().strftime("%Y-%m-%d")}.db')
if os.path.exists(today_db_path):
    st.warning(f"Ya existe una base de datos para hoy ({os.path.basename(today_db_path)})")
    if st.button("Eliminar base de datos existente"):
        try:
            os.remove(today_db_path)
            st.success("Base de datos eliminada correctamente")
        except Exception as e:
            st.error(f"Error al eliminar la base de datos: {str(e)}")

# Permitir al usuario seleccionar un archivo Excel desde la ruta predeterminada
excel_files = [f for f in os.listdir(default_excel_path) if f.endswith('.xlsx')]
selected_file = st.selectbox("Selecciona un archivo Excel", excel_files)

# Ruta completa del archivo Excel
uploaded_file_path = os.path.join(default_excel_path, selected_file)

# Ruta para almacenar la base de datos (hoy.db)
today_db_path = os.path.join(default_excel_path, f'{datetime.now().strftime("%Y-%m-%d")}.db')

# Botón para ejecutar la operación
if st.button("Ejecutar procesamiento de datos"):
    # Procesar el archivo Excel seleccionado
    all_data, invalid_data, total_records = process_excel_file(uploaded_file_path, default_mappings)
    
    # Crear la base de datos si no existe y luego insertar datos
    create_database(today_db_path)
    
    # Guardar los datos que se insertarán y los que no en DataFrames separados
    columns = ['Nombre', 'Cliente_Cuenta', 'Tipo_de_Dispositivo', 'IMEI', 'ICCID',
               'Fecha_de_Activacion', 'Fecha_de_Desactivacion', 'Hora_de_Ultimo_Mensaje',
               'Ultimo_Reporte', 'Vehiculo', 'Servicios', 'Grupo', 'Telefono', 'Origen', 'Fecha_Archivo']
    
    # Intentar insertar los datos y obtener los duplicados
    conn = sqlite3.connect(today_db_path)
    cursor = conn.cursor()
    
    # Lista para almacenar registros no insertados (duplicados)
    not_inserted = []
    inserted = []
    
    for record in all_data:
        try:
            cursor.execute('''
                INSERT INTO datos (
                    Nombre, Cliente_Cuenta, Tipo_de_Dispositivo, IMEI, ICCID,
                    Fecha_de_Activacion, Fecha_de_Desactivacion, Hora_de_Ultimo_Mensaje,
                    Ultimo_Reporte, Vehiculo, Servicios, Grupo, Telefono, Origen, Fecha_Archivo
                ) VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
            ''', record)
            inserted.append(record)
        except sqlite3.IntegrityError:
            not_inserted.append(record)
    
    conn.commit()
    conn.close()
    
    # Convertir a DataFrames
    df_inserted = pd.DataFrame(inserted, columns=columns)
    df_not_inserted = pd.DataFrame(not_inserted, columns=columns)
    
    # Mostrar resultados generales
    col1, col2, col3, col4 = st.columns(4)
    with col1:
        st.metric("Total de Registros", total_records)
    with col2:
        st.metric("Registros Insertados", len(inserted))
    with col3:
        st.metric("Registros No Insertados", len(not_inserted))
    with col4:
        st.metric("Registros Inválidos", len(invalid_data))

    # Mostrar registros no insertados
    if len(not_inserted) > 0:
        st.write("### Registros No Insertados en la Base de Datos")
        st.write("Los siguientes registros no fueron insertados por ser duplicados:")
        
        # Agregar filtros para los registros no insertados
        col1, col2 = st.columns(2)
        with col1:
            unique_clients = sorted(df_not_inserted['Cliente_Cuenta'].unique())
            selected_client = st.multiselect(
                'Filtrar por Cliente:',
                options=unique_clients,
                default=[]
            )
        
        with col2:
            unique_origins = sorted(df_not_inserted['Origen'].unique())
            selected_origin = st.multiselect(
                'Filtrar por Origen:',
                options=unique_origins,
                default=[]
            )
        
        # Aplicar filtros
        df_filtered = df_not_inserted.copy()
        if selected_client:
            df_filtered = df_filtered[df_filtered['Cliente_Cuenta'].isin(selected_client)]
        if selected_origin:
            df_filtered = df_filtered[df_filtered['Origen'].isin(selected_origin)]
        
        # Mostrar tabla con registros no insertados
        st.dataframe(df_filtered, use_container_width=True)
        
        # Botón para descargar registros no insertados
        csv = df_filtered.to_csv(index=False).encode('utf-8')
        st.download_button(
            label="Descargar registros no insertados como CSV",
            data=csv,
            file_name="registros_no_insertados.csv",
            mime='text/csv',
        )

    # Mostrar resumen por plataforma
    st.write("## Resumen por Plataforma")
    summary_data = []
    sheets = list(default_mappings.keys())
    for sheet in sheets:
        sheet_data = [record for record in all_data if record[-2] == sheet]
        total_sheet = len(sheet_data)
        percentage = (total_sheet / total_records * 100) if total_records > 0 else 0
        summary_data.append({
            "Plataforma": sheet,
            "Total Registros": total_sheet,
            "Porcentaje": f"{percentage:.1f}%"
        })
    
    # Crear DataFrame y mostrar resumen
    df_summary = pd.DataFrame(summary_data)
    st.dataframe(df_summary, use_container_width=True)
    
    # Mostrar gráfico de distribución
    st.write("### Distribución de Registros por Plataforma")
    df_summary['Porcentaje_Num'] = df_summary['Porcentaje'].str.rstrip('%').astype(float)
    chart_data = pd.DataFrame({
        'Plataforma': df_summary['Plataforma'],
        'Porcentaje': df_summary['Porcentaje_Num']
    })
    st.bar_chart(chart_data.set_index('Plataforma'))

    # Crear pestañas para cada plataforma
    tabs = st.tabs(sheets)
    
    for i, sheet in enumerate(sheets):
        with tabs[i]:
            st.write(f"## Análisis de {sheet}")
            
            # Obtener datos de esta pestaña
            sheet_data = [record for record in all_data if record[-2] == sheet]
            total_sheet = len(sheet_data)
            percentage = (total_sheet / total_records * 100) if total_records > 0 else 0
            
            # Mostrar resumen de la pestaña
            st.write("### Resumen de la Plataforma")
            col_summary1, col_summary2, col_summary3 = st.columns(3)
            with col_summary1:
                st.metric("Total Registros", total_sheet)
            with col_summary2:
                st.metric("Porcentaje del Total", f"{percentage:.1f}%")
            with col_summary3:
                st.metric("Campos Mapeados", sum(1 for v in default_mappings[sheet].values() if v is not None))
            
            # Crear columnas para estadísticas y mapeo
            col_stats, col_mapping = st.columns([3, 2])
            
            with col_mapping:
                st.write("### Mapeo de Campos")
                mapping_data = []
                for field, mapped_field in default_mappings[sheet].items():
                    status = "✅ Mapeado" if mapped_field else "❌ No mapeado"
                    mapping_data.append({
                        "Campo": field,
                        "Mapeo": mapped_field if mapped_field else "No disponible",
                        "Estado": status
                    })
                df_mapping = pd.DataFrame(mapping_data)
                st.dataframe(df_mapping, use_container_width=True)

            with col_stats:
                st.write("### Estadísticas de Datos")
                if total_sheet > 0:
                    field_stats = []
                    omitted_data = []
                    for idx, field in enumerate(['Nombre', 'Cliente_Cuenta', 'Tipo_de_Dispositivo', 'IMEI', 'ICCID',
                                              'Fecha_de_Activacion', 'Fecha_de_Desactivacion', 'Hora_de_Ultimo_Mensaje',
                                              'Ultimo_Reporte', 'Vehiculo', 'Servicios', 'Grupo', 'Telefono']):
                        non_empty = sum(1 for record in sheet_data if record[idx] is not None and str(record[idx]).strip())
                        empty = total_sheet - non_empty
                        percentage = (non_empty / total_sheet) * 100
                        
                        field_stats.append({
                            "Campo": field,
                            "Registros con Datos": non_empty,
                            "Registros sin Datos": empty,
                            "Porcentaje Completitud": f"{percentage:.1f}%"
                        })
                        
                        # Agregar a datos omitidos si hay registros vacíos
                        if empty > 0:
                            omitted_data.append({
                                "Campo": field,
                                "Registros Omitidos": empty,
                                "Porcentaje Omitido": f"{(empty/total_sheet)*100:.1f}%"
                            })
                    
                    df_stats = pd.DataFrame(field_stats)
                    st.dataframe(df_stats, use_container_width=True)
                    
                    # Mostrar datos omitidos
                    if omitted_data:
                        st.write("### Resumen de Datos Omitidos")
                        df_omitted = pd.DataFrame(omitted_data)
                        st.dataframe(df_omitted, use_container_width=True)
                        
                        # Gráfico de datos omitidos
                        st.write("### Visualización de Datos Omitidos")
                        df_omitted['Porcentaje_Num'] = df_omitted['Porcentaje Omitido'].str.rstrip('%').astype(float)
                        chart_data = pd.DataFrame({
                            'Campo': df_omitted['Campo'],
                            'Porcentaje': df_omitted['Porcentaje_Num']
                        })
                        st.bar_chart(chart_data.set_index('Campo'))

                        # Mostrar registros con datos omitidos
                        st.write("### Registros con Datos Omitidos")
                        st.write("Esta tabla muestra los registros que tienen uno o más campos sin datos:")
                        
                        # Convertir los datos de la pestaña a DataFrame
                        df_sheet = pd.DataFrame(sheet_data, columns=[
                            'Nombre', 'Cliente_Cuenta', 'Tipo_de_Dispositivo', 'IMEI', 'ICCID',
                            'Fecha_de_Activacion', 'Fecha_de_Desactivacion', 'Hora_de_Ultimo_Mensaje',
                            'Ultimo_Reporte', 'Vehiculo', 'Servicios', 'Grupo', 'Telefono', 'Origen', 'Fecha_Archivo'
                        ])
                        
                        # Función para verificar si una fila tiene datos omitidos
                        def has_missing_data(row):
                            return any(pd.isna(val) or str(val).strip() == '' for val in row)
                        
                        # Filtrar registros con datos omitidos
                        df_incomplete = df_sheet[df_sheet.apply(has_missing_data, axis=1)]
                        
                        if not df_incomplete.empty:
                            # Agregar columna que indica qué campos están vacíos
                            df_incomplete['Campos_Omitidos'] = df_incomplete.apply(
                                lambda row: ', '.join([col for col in df_incomplete.columns 
                                                     if pd.isna(row[col]) or str(row[col]).strip() == '']),
                                axis=1
                            )
                            
                            # Agregar filtros para la tabla de datos omitidos
                            col1, col2 = st.columns(2)
                            with col1:
                                if 'Cliente_Cuenta' in df_incomplete.columns:
                                    selected_client_incomplete = st.multiselect(
                                        'Filtrar por Cliente (Datos Omitidos):',
                                        options=sorted([x for x in df_incomplete['Cliente_Cuenta'].unique() 
                                                      if x is not None and str(x).strip()], key=str),
                                        default=[]
                                    )
                            with col2:
                                if 'Tipo_de_Dispositivo' in df_incomplete.columns:
                                    selected_device_incomplete = st.multiselect(
                                        'Filtrar por Tipo de Dispositivo (Datos Omitidos):',
                                        options=sorted([x for x in df_incomplete['Tipo_de_Dispositivo'].unique() 
                                                      if x is not None and str(x).strip()], key=str),
                                        default=[]
                                    )
                            
                            # Aplicar filtros a los datos omitidos
                            if selected_client_incomplete:
                                df_incomplete = df_incomplete[df_incomplete['Cliente_Cuenta'].isin(selected_client_incomplete)]
                            if selected_device_incomplete:
                                df_incomplete = df_incomplete[df_incomplete['Tipo_de_Dispositivo'].isin(selected_device_incomplete)]
                            
                            # Mostrar la tabla de datos omitidos con opción de descarga
                            st.dataframe(df_incomplete, use_container_width=True)
                            
                            # Botón para descargar datos omitidos
                            csv_incomplete = df_incomplete.to_csv(index=False).encode('utf-8')
                            st.download_button(
                                label=f"Descargar registros incompletos de {sheet} como CSV",
                                data=csv_incomplete,
                                file_name=f'{sheet}_registros_incompletos.csv',
                                mime='text/csv',
                            )
                        else:
                            st.info("No se encontraron registros con datos omitidos.")
                    
                    # Gráfico de completitud de datos
                    st.write("### Completitud de Datos por Campo")
                    df_stats['Porcentaje_Num'] = df_stats['Porcentaje Completitud'].str.rstrip('%').astype(float)
                    chart_data = pd.DataFrame({
                        'Campo': df_stats['Campo'],
                        'Porcentaje': df_stats['Porcentaje_Num']
                    })
                    st.bar_chart(chart_data.set_index('Campo'))
                else:
                    st.warning(f"No hay datos para analizar en la pestaña {sheet}")

            # Mostrar datos en tabla con filtros
            st.write("### Datos Detallados")
            if sheet_data:
                df = pd.DataFrame(sheet_data, columns=[
                    'Nombre', 'Cliente_Cuenta', 'Tipo_de_Dispositivo', 'IMEI', 'ICCID',
                    'Fecha_de_Activacion', 'Fecha_de_Desactivacion', 'Hora_de_Ultimo_Mensaje',
                    'Ultimo_Reporte', 'Vehiculo', 'Servicios', 'Grupo', 'Telefono', 'Origen', 'Fecha_Archivo'
                ])
                
                # Agregar filtros
                col1, col2 = st.columns(2)
                with col1:
                    if 'Cliente_Cuenta' in df.columns:
                        # Filtrar y manejar valores None
                        unique_clients = df['Cliente_Cuenta'].unique()
                        valid_clients = sorted([x for x in unique_clients if x is not None and str(x).strip()], key=str)
                        selected_client = st.multiselect(
                            'Filtrar por Cliente:',
                            options=valid_clients,
                            default=[]
                        )
                with col2:
                    if 'Tipo_de_Dispositivo' in df.columns:
                        # Filtrar y manejar valores None
                        unique_devices = df['Tipo_de_Dispositivo'].unique()
                        valid_devices = sorted([x for x in unique_devices if x is not None and str(x).strip()], key=str)
                        selected_device = st.multiselect(
                            'Filtrar por Tipo de Dispositivo:',
                            options=valid_devices,
                            default=[]
                        )
                
                # Aplicar filtros
                if selected_client:
                    df = df[df['Cliente_Cuenta'].isin(selected_client)]
                if selected_device:
                    df = df[df['Tipo_de_Dispositivo'].isin(selected_device)]
                
                # Mostrar datos filtrados
                st.dataframe(df, use_container_width=True)
                
                # Exportar datos
                csv = df.to_csv(index=False).encode('utf-8')
                st.download_button(
                    label=f"Descargar datos de {sheet} como CSV",
                    data=csv,
                    file_name=f'{sheet}_datos.csv',
                    mime='text/csv',
                )
            else:
                st.info(f"No hay datos disponibles para {sheet}")
